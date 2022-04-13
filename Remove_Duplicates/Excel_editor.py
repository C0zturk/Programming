#from openpyxl import Workbook
#from openpyxl import load_workbook
from typing import Counter, Pattern
from openpyxl import *
from openpyxl.cell import cell
from openpyxl.styles import *

workbook = load_workbook(filename="IPs_Excel.xlsx")
workbook.sheetnames
sheet = workbook.active
IP_List = []
# Checks value of A1 cell, if nothing is there it will add the string in brackets
column_counter=1
while sheet.cell(1, column_counter).value != None:
#create while loop here
    row_counter=1   
    #While loop adds DOB input above into cells by checking for 
    while sheet.cell(row_counter, column_counter).value != None:
        doesexist = False
        for IP in IP_List:
            if IP == sheet.cell(row_counter, column_counter).value and sheet.cell(row_counter, column_counter).value != "IP Address":
                doesexist = True
        if doesexist:
            sheet.cell(row_counter, column_counter).value = None
            row_counter = row_counter + 1
        else:
            IP_List.append(sheet.cell(row_counter, column_counter).value)
            row_counter = row_counter + 1
    column_counter = column_counter + 1




workbook.save(filename="IPS_Output.xlsx")

